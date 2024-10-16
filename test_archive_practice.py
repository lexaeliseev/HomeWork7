from conftest import TMP_DIR
from zipfile import ZipFile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook


def test_check_pdf(create_zip_file):
    with ZipFile(f"{TMP_DIR}/archive.zip") as zip_file:
        print(f"Список файлов в архиве: {zip_file.namelist()}")
        with zip_file.open("pdf_file.pdf") as pdf_file:
            reader_pdf = PdfReader(pdf_file)
            count_pages = len(reader_pdf.pages)
            print(f"Количество страниц в файле PDF: {count_pages}")
            assert count_pages == 256

            check_text = reader_pdf.pages[255].extract_text().strip()
            print(f"Текст со страницы {count_pages}: {check_text}")
            assert "Mathematics is beautiful—and it can be fun and exciting as well as practical." in check_text


def test_check_xlsx(create_zip_file):
    with ZipFile(f"{TMP_DIR}/archive.zip") as zip_file:
        with zip_file.open("import_ou.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)

            count_pages = len(workbook.sheetnames)
            print(f"Количество листов в файле XLSX: {count_pages}")
            assert count_pages == 1

            sheet = workbook.active
            print(f"Название активного листа в файле XLSX: {sheet.title}")
            assert sheet.title == "TemplateImportOU"

            cell_value = sheet['C1'].value
            print(f"Значение ячейки: {cell_value}")
            assert cell_value == "Название"


def test_check_csv(create_zip_file):
    with ZipFile(f"{TMP_DIR}/archive.zip") as zip_file:
        with zip_file.open("csv.csv") as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            reader_csv = list(csv.reader(content.splitlines()))

            first_row = reader_csv[0]
            count_columns = len(first_row)
            print(f"Количество колонок в файле: {count_columns}")
            assert count_columns == 3

            value = "Внешний идентификатор для импорта,Вышестоящий отдел,Название"
            list_value = value.split(",")
            print(f"Список заголовков в файле: {first_row}")
            assert first_row == list_value

            count_row = sum(1 for _ in reader_csv)
            print(f"Количество строк в файле: {count_row}")
            assert count_row == 7
