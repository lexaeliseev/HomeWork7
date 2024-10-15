from openpyxl import load_workbook

workbook = load_workbook("download_file/fake_and_links.2024-09-26 10_48_19.216625 (1).xlsx")

sheets = workbook.sheetnames  # список листов
for sheet in sheets:
    print(sheet)

sheet = workbook.active  # получаем активный лист

print(sheet["A1"].value)  # печатаем значение ячейки A1
print(sheet["B1"].value)  # печатаем значение ячейки B1

sheet2 = workbook["Фейки"]  # получаем другой лист(в скобках название листа)
print(sheet2["A1"].value)  # печатаем значение ячейки A1 листа sheet2

workbook.active = 0  # делаем 3 лист активным

cell = sheet["C1"]
print('Строка: ' + str(cell.row))       # номер строки
print('Столбец: ' + str(cell.column))   # номер столбца
print('Ячейка: ' + cell.coordinate)     # местоположение ячейки

cell = sheet.cell(row=1, column=4)      # получаем значение из 2 строки и 3 столбца
print(cell.value)

for cell in sheet['E']:  # все значения из указанной колонки
    print(cell.value)



